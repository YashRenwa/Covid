import openpyxl as xl
from datetime import date

cases = xl.load_workbook('Covid-India.xlsx')
recover = xl.load_workbook('time_series_covid_19_recovered.xlsx')
sheet_recover = recover['Worksheet']
sheet_cases = cases['Worksheet']
number_row_recover = sheet_recover.max_row
number_column_recover = sheet_recover.max_column
number_row_cases = sheet_cases.max_row
number_column_cases = sheet_cases.max_column
averageDays = []

for i in range(2, number_row_recover + 1):
    cell = sheet_cases.cell(1, i)
    list_confirmed_patients = {}
    list_recovered_patients = {}
    list_days = {}
    k = 0
    for j in range(i, number_row_cases + 1):
        if j % 2 != 0:
            cell_cases = sheet_cases.cell(j, i)
            cell_date = sheet_cases.cell(j, 1)
            if cell_cases.value != 0:
                for l in range(0, cell_cases.value):
                    if int(cell_date.value[1]) != 0:
                        list_confirmed_patients[k] = cell_date.value[1:]
                        k += 1
                    else:
                        list_confirmed_patients[k] = cell_date.value[2:]
                        k += 1

    k = 0
    for j in range(i + 3, number_column_recover + 1):
        cell_recover = sheet_recover.cell(i, j)
        cell_date = sheet_recover.cell(1, j)
        if i == 2:
            cell_date.value += '20'
        if cell_recover.value != 0:
            for l in range(0, cell_recover.value):
                list_recovered_patients[k] = cell_date.value
                k += 1

    print(list_confirmed_patients)
    print(list_recovered_patients)

    listofdays = []
    for j in range(0, len(list_recovered_patients)):
        if list_recovered_patients[j][3] != '/':
            f_date = date(int(list_confirmed_patients[j][5:]), int(list_confirmed_patients[j][0]),
                          int(list_confirmed_patients[j][2:4]))
            l_date = date(int(list_recovered_patients[j][5:]), int(list_recovered_patients[j][0]),
                          int(list_recovered_patients[j][2:4]))
            delta = l_date - f_date
            listofdays.append(delta.days)
        else:
            f_date = date(int(list_confirmed_patients[j][5:]), int(list_confirmed_patients[j][0]),
                          int(list_confirmed_patients[j][2:4]))
            l_date = date(int(list_recovered_patients[j][4:]), int(list_recovered_patients[j][0]),
                          int(list_recovered_patients[j][2]))
            delta = l_date - f_date
            listofdays.append(delta.days)
    print(listofdays)
    average = 0
    for j in range(0, len(listofdays)):
        average += listofdays[j]
    if len(listofdays) != 0:
        average = average / len(listofdays)
    else:
        average = 0
    averageDays.append(average)

print(averageDays)
print(max(averageDays))
index = averageDays.index(max(averageDays))
country = sheet_recover.cell(index + 1, 2)
print(country.value)
